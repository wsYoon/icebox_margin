import sys
from PyQt5 import QtCore,QtGui,QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import QSize, Qt
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtGui import QFont
import math
import os
import openpyxl
import xlrd
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import random
import csv
import re

#UI파일 연결
#단, UI파일은 Python 코드 파일과 같은 디렉토리에 위치해야한다.
form_class = uic.loadUiType("babs_margin.ui")[0]

class WindowClass(QMainWindow, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("밥스토랑 마진계산기")
        self.pushButton_calc.clicked.connect(self.button_margincal)          #마진계산 버튼
        #self.pushButton_calc.setFont(QFont("나눔고딕", 11))    #폰트 변경
        # 엑셀파일 찾아보기 버튼 함수 정의
        self.pushButton_search_mainlist.clicked.connect(self.fileopen_mainlist)
        self.pushButton_search_easyadminlist.clicked.connect(self.fileopen_easyadminlist)
        self.pushButton_search_lottelist.clicked.connect(self.fileopen_lottelist)
        self.pushButton_search_snack365list.clicked.connect(self.fileopen_snack365list)
        self.pushButton_search_onegalist.clicked.connect(self.fileopen_onegalist)
        
        # 상온체크박스 동작
        self.checkbox_S = self.findChild(QCheckBox, "checkBox_S")
        # 체크박스의 상태 변화를 감지하기 위한 시그널 연결
        self.checkbox_S.stateChanged.connect(self.checkboxStateChanged)
        self.lineEdit_snack365list.setEnabled(False)
        self.lineEdit_onegalist.setEnabled(False)
        self.pushButton_search_snack365list.setEnabled(False)
        self.pushButton_search_onegalist.setEnabled(False)

        #운임 변환 & 박스비 & 드라이아이스
        self.tableWidget_costchange.setRowCount(13)
        self.tableWidget_costchange.setColumnCount(5)
        self.tableWidget_costchange.setColumnWidth(0,90)
        self.tableWidget_costchange.setColumnWidth(1,110)
        self.tableWidget_costchange.setColumnWidth(2,70)
        self.tableWidget_costchange.setColumnWidth(3,70)
        self.tableWidget_costchange.setColumnWidth(4,110)

        self.load_csv_costchange()
        self.pushButton_costchange.clicked.connect(self.save_csv_costchange)

        #주문거래처 변경
        self.tableWidget_namechange.setRowCount(5)
        self.tableWidget_namechange.setColumnCount(2)
        self.tableWidget_namechange.setColumnWidth(0,200)
        self.tableWidget_namechange.setColumnWidth(1,200)

        self.load_csv_namechange()
        self.pushButton_namechange.clicked.connect(self.save_csv_namechange)
        


    def button_margincal(self):
        try:
            # 엑셀파일 Dataframe으로 불러오기
            mainlist_wb = load_workbook(filename_mainlist[0])    # 엑셀 파일 불러오기
            mainlist_sheet = mainlist_wb.active     #시트 선택
            mainlist_sheet.delete_rows(1)
            mainlist = pd.DataFrame(mainlist_sheet.values)           # 데이터프레임으로 변환
            mainlist.columns = mainlist.iloc[0]       # 첫 번째 행을 열 인덱스로 설정
            mainlist = mainlist.iloc[1:]        # 첫 번째 행 삭제
            mainlist = mainlist.sort_values(by=['배송일자', '주문거래처'])
            mainlist.reset_index(drop=True, inplace=True)   # 행인덱스 초기화
            easyadminlist = pd.read_excel(filename_easyadminlist[0],engine='xlrd')
            easyadminlist = easyadminlist.drop(easyadminlist.index[-1]) #마지막줄 '합계' 행 삭제
            easyadminlist['배송일'] = easyadminlist['배송일'].apply(lambda x: x.to_pydatetime().strftime("%Y%m%d"))   #날짜형식 yyyymmdd로 변경
            lottelist = pd.read_excel(filename_lottelist[0],engine='openpyxl')
            lottelist['운송장번호'] = lottelist['운송장번호'].str.replace('-','')
            #lottelist = lottelist[lottelist['상품명'].str.contains('(밥스)|추가송장')]
            lottelist.reset_index(drop=True, inplace=True)   # 행인덱스 초기화
            checkbox_state_상온 = self.checkbox_S.isChecked()
            if checkbox_state_상온 == True:
                snack365list = pd.read_excel(filename_snack365list[0],engine='openpyxl')
                onegalist = pd.read_excel(filename_onegalist[0],engine='openpyxl')

            #######임시 경로지정#######
            # mainfile = 'D:\카카오톡다운로드\거래처별거래원장 납기일기준(12.01-12.31) _ 상온.xlsx'
            # easyadminfile = 'D:\카카오톡다운로드\이지어드민 밥스토랑냉동출고내역 (11.30-12.29) 상온포함.xls'
            # lottefile = r"D:\카카오톡다운로드\12월 롯데택배 출고 전체리스트 (3).xlsx"
            # snack365file = 'D:\카카오톡다운로드\범용상품리스트_이오스_(상온).xlsx'
            # onegafile = 'D:\카카오톡다운로드\원가표 (7).xlsx'
            # mainlist_wb = load_workbook(mainfile)    # 엑셀 파일 불러오기
            # mainlist_sheet = mainlist_wb.active     #시트 선택
            # mainlist_sheet.delete_rows(1)
            # mainlist = pd.DataFrame(mainlist_sheet.values)           # 데이터프레임으로 변환
            # mainlist.columns = mainlist.iloc[0]       # 첫 번째 행을 열 인덱스로 설정
            # mainlist = mainlist.iloc[1:]        # 첫 번째 행 삭제
            # mainlist = mainlist.sort_values(by=['배송일자', '주문거래처'])
            # mainlist.reset_index(drop=True, inplace=True)   # 행인덱스 초기화
            # easyadminlist = pd.read_excel(easyadminfile,engine='xlrd')
            # easyadminlist = easyadminlist.drop(easyadminlist.index[-1]) #마지막줄 '합계' 행 삭제
            # easyadminlist['배송일'] = easyadminlist['배송일'].apply(lambda x: x.to_pydatetime().strftime("%Y%m%d"))   #날짜형식 yyyymmdd로 변경
            # lottelist = pd.read_excel(lottefile,engine='openpyxl')
            # lottelist['운송장번호'] = lottelist['운송장번호'].str.replace('-','')
            # #lottelist = lottelist[lottelist['상품명'].str.contains('(밥스)|추가송장')]
            # lottelist.reset_index(drop=True, inplace=True)   # 행인덱스 초기화
            # checkbox_state_상온 = self.checkbox_S.isChecked()
            # if checkbox_state_상온 == True:
            #     snack365list = pd.read_excel(snack365file,engine='openpyxl')
            #     onegalist = pd.read_excel(onegafile,engine='openpyxl')
            ##############################

            #거래원장의 날짜 리스트 (최종결과 시트 출력 시에 사용)
            dates = mainlist['배송일자'].dropna().unique().tolist()
            # 각 날짜에 '_상온'을 붙여서 새로운 항목 생성 후 원래 리스트에 추가
            new_dates=[]
            for date in dates:
                new_dates.append(date)
                new_dates.append(date + '_상온')
                new_dates.append(date + '_바바리아')
            dates = new_dates

            #mainlist의 소계, 총계 행 삭제(품목코드 없는 행 삭제)
            mainlist = mainlist.dropna(subset=['품목코드'])
            mainlist.reset_index(drop=True, inplace=True)   # 행인덱스 초기화
            
            # 운임요금 표의 데이터프레임화
            data = []
            columns = []
            for col in range(self.tableWidget_costchange.columnCount()):
                header_item = self.tableWidget_costchange.horizontalHeaderItem(col)
                column_name = header_item.text() if header_item and header_item.text() else f'Column_{col}'
                columns.append(column_name)
            for row in range(self.tableWidget_costchange.rowCount()):
                row_data = []
                for col in range(self.tableWidget_costchange.columnCount()):
                    item = self.tableWidget_costchange.item(row, col)
                    cell_value = item.text() if item and item.text() else ''
                    row_data.append(cell_value)
                data.append(row_data)
            costchange = pd.DataFrame(data, columns=columns)
            costchange['기본운임'] = costchange['기본운임'].str.replace('\ufeff', '').astype(int)
            costchange[['변환택배비','박스값','드라이아이스']] = costchange[['변환택배비','박스값','드라이아이스']].astype(int)

            # 주문거래처 변경 표의 데이터프레임화
            data = []
            columns = []
            for col in range(self.tableWidget_namechange.columnCount()):
                header_item = self.tableWidget_namechange.horizontalHeaderItem(col)
                column_name = header_item.text() if header_item and header_item.text() else f'Column_{col}'
                columns.append(column_name)
            for row in range(self.tableWidget_namechange.rowCount()):
                row_data = []
                for col in range(self.tableWidget_namechange.columnCount()):
                    item = self.tableWidget_namechange.item(row, col)
                    cell_value = item.text() if item and item.text() else ''
                    row_data.append(cell_value)
                data.append(row_data)
            namechange = pd.DataFrame(data, columns=columns)
            #주문거래처, 수령자이름 통일하기
            for n in range(len(namechange)):
                namechange['변경 전'][n] = namechange['변경 전'][n].replace('\ufeff', '')
                namechange['변경 후'][n] = namechange['변경 후'][n].replace('\ufeff', '')
                mainlist['주문거래처'] = mainlist['주문거래처'].str.replace(namechange['변경 전'][n],namechange['변경 후'][n])
                easyadminlist['수령자이름'] = easyadminlist['수령자이름'].str.replace(namechange['변경 전'][n],namechange['변경 후'][n])



            # 최종 결과파일 데이터프레임 생성
            result_list = pd.DataFrame(columns=['주문거래처','배송일자','최종운임','주문금액','택배비','수수료(8%)','최종마진금액'])

            조회한주문거래처 = ''
            조회한날짜 = ''
            조회한코드 = ''

            드라이아이스_냉동 = 0
            택배비_냉동 = 0
            박스값_냉동 = 0
            드라이아이스_상온 = 0
            택배비_상온 = 0
            박스값_상온 = 0

            # 거래원장(mainlist) 항목을 하나씩 조회하면서 송장번호(이지어드민)을 통해 택배운임(롯데택배)을 검색하여 total_list 작성
            for i in range(len(mainlist)):
            #result_list 채우기
                #이미 조회한 날짜의 주문거래처인 경우 pass
                print(i+1, mainlist['코드'][i], mainlist['주문거래처'][i], mainlist['배송일자'][i])
                if (mainlist['코드'][i]==조회한코드) & (mainlist['주문거래처'][i] == 조회한주문거래처) & (mainlist['배송일자'][i] == 조회한날짜):
                    continue
                조회한주문거래처 = mainlist['주문거래처'][i]
                조회한날짜 = mainlist['배송일자'][i]
                조회한코드 = mainlist['코드'][i]

                # 조회된 행과 동일한 주문거래처, 발송일인 상품끼리 묶기
                original_selected_list = mainlist.loc[(mainlist['코드'] == mainlist['코드'][i]) &(mainlist['주문거래처'] == mainlist['주문거래처'][i]) & (mainlist['배송일자'] == mainlist['배송일자'][i]), '품명'].tolist()
                selected_list = [item.replace('NEW_', '') for item in original_selected_list]
                selected_list = [item.replace('BOBs ', '') for item in original_selected_list]

                #택배비를 제외하고 아무 상품이나 하나 뽑기 (이지어드민과 대조해서 송장번호를 찾을 것임)
                selected_item = random.choice([item for item in selected_list if item != '택배비'])
                
                #******스낵365 상품(바바리아 포함)******
                #if selected_item in snack365list['상품명'].values:
                if checkbox_state_상온 == True:
                    #주문금액, 상품원가
                    주문금액 = 0
                    상품원가 = 0
                    바바리아마진차 = 0    # 바바리아 상품 마진은 "판매가*11% + 3000"으로 고정이므로 주문금액에서 뺄 마진차인 "판매가*0.89 -3000" 생성
                    for item in original_selected_list:
                        if item != '택배비':
                            if '바바리아' in item:  #바바리아 제품
                                주문금액 = mainlist.loc[(mainlist['배송일자'] == mainlist['배송일자'][i]) & (mainlist['주문거래처'] == mainlist['주문거래처'][i]) & (mainlist['품명'] != '택배비'), '합계금액'].sum()
                                바바리아마진차 += 주문금액*0.89 - 3000
                            else:   #스낵365 제품
                                주문금액 += snack365list.loc[snack365list['상품명'].str.contains(re.sub(r'\([^)]*\)', '', item)), '이오스'].values[0] * mainlist.loc[(mainlist['품명'] == item) & (mainlist['배송일자'] == mainlist['배송일자'][i]) & (mainlist['주문거래처'] == mainlist['주문거래처'][i]), '배송량'].values[0]
                                상품원가 += onegalist.loc[onegalist['상품명'].str.contains(re.sub(r'\([^)]*\)', '', item)),'박스금액'].values[0] * mainlist.loc[(mainlist['품명'] == item) & (mainlist['배송일자'] == mainlist['배송일자'][i]) & (mainlist['주문거래처'] == mainlist['주문거래처'][i]), '배송량'].values[0]

                    #송장번호
                    songjang =[]
                    for item in original_selected_list:
                        if item != '택배비':
                            #주문금액 += mainlist.loc[(mainlist['품명'] == item) & (mainlist['배송일자'] == mainlist['배송일자'][i]) & (mainlist['주문거래처'] == mainlist['주문거래처'][i])), '합계금액'].values[0]
                            #이지어드민에서 배송일, 상품명(추가송장 포함), 주문거래처가 일치하는 송장번호를 리스트화
                            if '(' in item and item.endswith(')'):
                                item = item.rsplit(' (', 1)[0]
                            songjang_item_list = easyadminlist[(easyadminlist['판매처 상품명'].str.contains(item)) & (easyadminlist['배송일'] == mainlist['배송일자'][i]) & (easyadminlist['수령자이름']==mainlist['주문거래처'][i])]['송장번호'].tolist()
                            songjang.extend(songjang_item for songjang_item in songjang_item_list if songjang_item not in songjang)
                    #추가송장 확인
                    songjang_추가 = easyadminlist[(easyadminlist['상품명'].str.contains('추가송장')) & (easyadminlist['상품명'].str.contains('밥스') == False) & (easyadminlist['배송일'] == mainlist['배송일자'][i]) & (easyadminlist['수령자이름']==mainlist['주문거래처'][i])]['송장번호'].tolist()
                    songjang.extend(songjang_추가)
                    songjang = [int(item) for item in songjang] #int로 변경

                    운임합계 = 0
                    택배비 = 0

                    운임합계 = 0
                    for item_songjang in songjang:
                        롯데택배운임 = lottelist.loc[lottelist['운송장번호']==str(item_songjang), '운임합계'].values[0]
                        변환택배비 = costchange.loc[costchange['기본운임']==int(롯데택배운임), '변환택배비'].values[0]
                        if any('(냉동)' in item for item in map(str, original_selected_list)):
                            박스값 = costchange.loc[costchange['기본운임']==int(롯데택배운임), '박스값'].values[0]
                            박스값_냉동 += 박스값
                            드라이아이스 = costchange.loc[costchange['기본운임']==int(롯데택배운임), '드라이아이스'].values[0]
                            드라이아이스_냉동 += 드라이아이스 #드라이아이스 금액 총합계
                            운임합계 += (변환택배비+박스값+드라이아이스)
                        else:
                            운임합계 += 변환택배비

                    if '바바리아' in item:
                        result_list = result_list.append(
                                {'주문거래처':mainlist['주문거래처'][i],
                                '최종운임':운임합계,
                                '택배비':택배비,
                                '배송일자':mainlist['배송일자'][i]+'_바바리아',
                                '주문금액':주문금액,
                                '수수료(8%)':0,
                                '최종마진금액':주문금액-운임합계-바바리아마진차},
                                ignore_index=True
                        )
                    else:
                        result_list = result_list.append(
                                {'주문거래처':mainlist['주문거래처'][i],
                                '최종운임':운임합계,
                                '택배비':택배비,
                                '배송일자':mainlist['배송일자'][i]+'_상온',
                                '주문금액':주문금액,
                                '수수료(8%)':0,
                                '최종마진금액':주문금액-상품원가-운임합계},
                                ignore_index=True
                        )
                

                #******밥스토랑 상품******
                else:                
                    #selected_item에 택배비가 있으면 result_list에 택배비 추가
                    택배비 = 0
                    택배비 = 10000 * selected_list.count('택배비')
                    택배비_냉동 += 택배비   #택배비 금액 총 합계

                    #주문금액
                    주문금액 = mainlist.loc[(mainlist['배송일자'] == mainlist['배송일자'][i]) & (mainlist['주문거래처'] == mainlist['주문거래처'][i]) & (mainlist['품명'] != '택배비'), '합계금액'].sum()

                    #송장번호
                    songjang =[]
                    for item in original_selected_list:
                        if item != '택배비':
                            #주문금액 += mainlist.loc[(mainlist['품명'] == item) & (mainlist['배송일자'] == mainlist['배송일자'][i]) & (mainlist['주문거래처'] == mainlist['주문거래처'][i])), '합계금액'].values[0]
                            #이지어드민에서 배송일, 상품명(추가송장 포함), 주문거래처가 일치하는 송장번호를 리스트화
                            songjang_item_list = easyadminlist[(easyadminlist['판매처 상품명'].str.contains(item)) & (easyadminlist['배송일'] == mainlist['배송일자'][i]) & (easyadminlist['수령자이름']==mainlist['주문거래처'][i])]['송장번호'].tolist()
                            songjang.extend(songjang_item for songjang_item in songjang_item_list if songjang_item not in songjang)
                    #추가송장 확인
                    songjang_추가 = easyadminlist[(easyadminlist['상품명'].str.contains('추가송장')) & (easyadminlist['상품명'].str.contains('밥스')) & (easyadminlist['배송일'] == mainlist['배송일자'][i]) & (easyadminlist['수령자이름']==mainlist['주문거래처'][i])]['송장번호'].tolist()
                    songjang.extend(songjang_추가)
                    songjang = [int(item) for item in songjang] #int로 변경

                    운임합계 = 0
                    for item_songjang in songjang:
                        롯데택배운임 = lottelist.loc[lottelist['운송장번호']==str(item_songjang), '운임합계'].values[0]
                        변환택배비 = costchange.loc[costchange['기본운임']==int(롯데택배운임), '변환택배비'].values[0]
                        박스값 = costchange.loc[costchange['기본운임']==int(롯데택배운임), '박스값'].values[0]
                        박스값_냉동 += 박스값
                        드라이아이스 = costchange.loc[costchange['기본운임']==int(롯데택배운임), '드라이아이스'].values[0]
                        드라이아이스_냉동 += 드라이아이스 #드라이아이스 금액 총합계
                        운임합계 += (변환택배비+박스값+드라이아이스)

                    result_list = result_list.append(
                            {'주문거래처':mainlist['주문거래처'][i],
                            '최종운임':운임합계,
                            '택배비':택배비,
                            '배송일자':mainlist['배송일자'][i],
                            '주문금액':주문금액,
                            '수수료(8%)':주문금액*0.08,
                            '최종마진금액':주문금액*0.08 + 택배비 - 운임합계},
                            ignore_index=True
                    )

            #최종결과 변수 초기화
            판매금액_냉동 = 0
            판매금액_상온 = 0
            마진_냉동 = 0
            마진_상온 = 0
            

            #결과 엑셀파일 작성
            workbook = openpyxl.Workbook()
            for sht in dates:
                sheet = workbook.create_sheet(title = sht)  #날짜를 시트이름으로 하는 시트 생성

                sheet['A1'] = '수령자이름'
                sheet['B1'] = '최종운임'
                sheet['C1'] = '택배비'
                sheet['D1'] = '주문금액'
                sheet['E1'] = '수수료(8%)'
                sheet['F1'] = '최종마진금액'

                #해당 날짜와 일치하는 행만 별도의 데이터프레임 생성
                filtered_df = result_list[result_list['배송일자']==sht]
                filtered_df.reset_index(drop=True, inplace=True)   # 행인덱스 초기화
                if filtered_df.empty:
                    workbook.remove_sheet(workbook[sht])
                    continue

                최종운임_합계 = 0
                택배비_합계 = 0
                주문금액_합계 = 0
                수수료_합계 = 0
                최종마진금액_합계 = 0
                for k in range(1, len(filtered_df)+1):
                    sheet[f'A{k+1}'] = filtered_df['주문거래처'][k-1]
                    sheet[f'B{k+1}'] = filtered_df['최종운임'][k-1]
                    sheet[f'C{k+1}'] = filtered_df['택배비'][k-1]
                    sheet[f'D{k+1}'] = filtered_df['주문금액'][k-1]
                    sheet[f'E{k+1}'] = filtered_df['수수료(8%)'][k-1]
                    sheet[f'F{k+1}'] = filtered_df['최종마진금액'][k-1]
                    최종운임_합계 += filtered_df['최종운임'][k-1]
                    택배비_합계 += filtered_df['택배비'][k-1]
                    주문금액_합계 += filtered_df['주문금액'][k-1]
                    수수료_합계 += filtered_df['수수료(8%)'][k-1]
                    최종마진금액_합계 += filtered_df['최종마진금액'][k-1]

                    if ('상온' in sht) or ('바바리아' in sht):
                        판매금액_상온 += filtered_df['주문금액'][k-1]
                        마진_상온 += filtered_df['최종마진금액'][k-1]
                    else :
                        판매금액_냉동 += filtered_df['주문금액'][k-1]
                        마진_냉동 += filtered_df['최종마진금액'][k-1]
                
                sheet[f'A{k+2}'] = '합계'
                sheet[f'B{k+2}'] = 최종운임_합계
                sheet[f'C{k+2}'] = 택배비_합계
                sheet[f'D{k+2}'] = 주문금액_합계
                sheet[f'E{k+2}'] = 수수료_합계
                sheet[f'F{k+2}'] = 최종마진금액_합계

                sheet.column_dimensions['A'].width = 30
                sheet.column_dimensions['E'].width = 15
                sheet.column_dimensions['F'].width = 15

            workbook.remove_sheet(workbook['Sheet'])
            

            #GUI 최종 결과
            self.lineEdit_sell_N.setText(format(판매금액_냉동,','))
            self.lineEdit_margin_N.setText('{:,.1f}'.format(마진_냉동))
            self.lineEdit_sell_S.setText(format(판매금액_상온,','))
            self.lineEdit_margin_S.setText('{:,.1f}'.format(마진_상온))
            self.lineEdit_dryice_N.setText(format(드라이아이스_냉동,','))
            self.lineEdit_dryice_S.setText(format(드라이아이스_상온,','))
            self.lineEdit_icebox_N.setText(format(박스값_냉동,','))
            self.lineEdit_icebox_S.setText(format(박스값_상온,','))
            self.lineEdit_TB_N.setText(format(택배비_냉동,','))
            self.lineEdit_TB_S.setText(format(택배비_상온,','))

            #엑셀파일 첫페이지에 최종 결과 기록
            try:    #냉동
                workbook[dates[0]].column_dimensions['H'].width = 15
                workbook[dates[0]].column_dimensions['I'].width = 15
                workbook[dates[0]].column_dimensions['J'].width = 15

                workbook[dates[0]]['I1'] = '냉동'
                workbook[dates[0]]['J1'] = '상온'
                workbook[dates[0]]['H2'] = '총판매금액'
                workbook[dates[0]]['H3'] = '드라이아이스'
                workbook[dates[0]]['H4'] = '아이스박스'
                workbook[dates[0]]['H5'] = '택배비'
                workbook[dates[0]]['H6'] = '마진'
                workbook[dates[0]]['I2'] = format(판매금액_냉동,',')
                workbook[dates[0]]['I3'] = format(드라이아이스_냉동,',')
                workbook[dates[0]]['I4'] = format(박스값_냉동,',')
                workbook[dates[0]]['I5'] = format(택배비_냉동,',')
                workbook[dates[0]]['I6'] = '{:,.1f}'.format(마진_냉동)
                workbook[dates[0]]['J2'] = format(판매금액_상온,',')
                workbook[dates[0]]['J6'] = '{:,.1f}'.format(마진_상온)
            except: #상온
                workbook[dates[1]].column_dimensions['H'].width = 15
                workbook[dates[1]].column_dimensions['I'].width = 15
                workbook[dates[1]].column_dimensions['J'].width = 15

                workbook[dates[1]]['I1'] = '냉동'
                workbook[dates[1]]['J1'] = '상온'
                workbook[dates[1]]['H2'] = '총판매금액'
                workbook[dates[1]]['H3'] = '드라이아이스'
                workbook[dates[1]]['H4'] = '아이스박스'
                workbook[dates[1]]['H5'] = '택배비'
                workbook[dates[1]]['H6'] = '마진'
                workbook[dates[1]]['I2'] = format(판매금액_냉동,',')
                workbook[dates[1]]['I3'] = format(드라이아이스_냉동,',')
                workbook[dates[1]]['I4'] = format(박스값_냉동,',')
                workbook[dates[1]]['I5'] = format(택배비_냉동,',')
                workbook[dates[1]]['I6'] = '{:,.1f}'.format(마진_냉동)
                workbook[dates[1]]['J2'] = format(판매금액_상온,',')
                workbook[dates[1]]['J6'] = '{:,.1f}'.format(마진_상온)


            workbook.save('밥스마진계산결과.xlsx')

        except Exception as e:
            QMessageBox.critical(None, '에러', f'주문거래처: {mainlist["주문거래처"][i]}, 배송일자: {mainlist["배송일자"][i]}, 송장번호: {item_songjang}')

    # 엑셀파일 찾아보기 기능
    def fileopen_mainlist(self):
        global filename_mainlist
        filename_mainlist = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_mainlist.setText(filename_mainlist[0])

    def fileopen_easyadminlist(self):
        global filename_easyadminlist
        filename_easyadminlist = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_easyadminlist.setText(filename_easyadminlist[0])

    def fileopen_lottelist(self):
        global filename_lottelist
        filename_lottelist = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_lottelist.setText(filename_lottelist[0])

    def fileopen_snack365list(self):
        global filename_snack365list
        filename_snack365list = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_snack365list.setText(filename_snack365list[0])

    def fileopen_onegalist(self):
        global filename_onegalist
        filename_onegalist = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_onegalist.setText(filename_onegalist[0])

    def load_csv_costchange(self):
        try:
            with open('운임비별 변환택배비.csv', 'r', encoding='utf-8') as f:
                rdr = csv.reader(f)
                i = 0
                for line in rdr:
                    self.tableWidget_costchange.setItem(i, 0, QTableWidgetItem(line[0]))
                    self.tableWidget_costchange.setItem(i, 1, QTableWidgetItem(line[1]))
                    self.tableWidget_costchange.setItem(i, 2, QTableWidgetItem(line[2]))
                    self.tableWidget_costchange.setItem(i, 3, QTableWidgetItem(line[3]))
                    self.tableWidget_costchange.setItem(i, 4, QTableWidgetItem(line[4]))
                    # 가운데 정렬
                    for col in range(self.tableWidget_costchange.columnCount()):
                        item = self.tableWidget_costchange.item(i, col)
                        item.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget_costchange.setRowHeight(i, 25) # 행 높이 수정
                    i += 1
        except Exception as e:
            print(f"Error loading CSV data: {e}")

    def save_csv_costchange(self):
        try:
            with open('운임비별 변환택배비.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                for row in range(self.tableWidget_costchange.rowCount()):
                    row_data = []
                    for col in range(self.tableWidget_costchange.columnCount()):
                        item = self.tableWidget_costchange.item(row, col)
                        if item is not None:
                            row_data.append(item.text())
                        else:
                            row_data.append('')  # Handle the case where the item is None (empty cell)

                    writer.writerow(row_data)

            print("CSV file saved successfully.")
        except Exception as e:
            print(f"Error saving CSV data: {e}")

    def load_csv_namechange(self):
        try:
            with open('주문거래처 변경.csv', 'r', encoding='utf-8') as f:
                rdr = csv.reader(f)
                i = 0
                for line in rdr:
                    self.tableWidget_namechange.setItem(i, 0, QTableWidgetItem(line[0]))
                    self.tableWidget_namechange.setItem(i, 1, QTableWidgetItem(line[1]))
                    # 가운데 정렬
                    for col in range(self.tableWidget_namechange.columnCount()):
                        item = self.tableWidget_namechange.item(i, col)
                        item.setTextAlignment(Qt.AlignCenter)
                    i += 1
        except Exception as e:
            print(f"Error loading CSV data: {e}")

    def save_csv_namechange(self):
        try:
            with open('주문거래처 변경.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                for row in range(self.tableWidget_namechange.rowCount()):
                    row_data = []
                    for col in range(self.tableWidget_namechange.columnCount()):
                        item = self.tableWidget_namechange.item(row, col)
                        if item is not None:
                            row_data.append(item.text())
                        else:
                            row_data.append('')  # Handle the case where the item is None (empty cell)

                    writer.writerow(row_data)

            print("CSV file saved successfully.")
        except Exception as e:
            print(f"Error saving CSV data: {e}")

    def checkboxStateChanged(self, state):
        #체크박스 상태에 따라 동작을 수행
        if state == 2:  #체크되었을 때 : 상온
            self.lineEdit_snack365list.setEnabled(True)
            self.lineEdit_onegalist.setEnabled(True)
            self.pushButton_search_snack365list.setEnabled(True)
            self.pushButton_search_onegalist.setEnabled(True)
        else:   # 체크되지 않았을 때 : 냉동
            self.lineEdit_snack365list.setEnabled(False)
            self.lineEdit_onegalist.setEnabled(False)
            self.pushButton_search_snack365list.setEnabled(False)
            self.pushButton_search_onegalist.setEnabled(False)



if __name__ == "__main__" :
    #QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv) 
    #WindowClass의 인스턴스 생성
    myWindow = WindowClass() 
    #프로그램 화면을 보여주는 코드
    myWindow.show()
    #프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    app.exec_()